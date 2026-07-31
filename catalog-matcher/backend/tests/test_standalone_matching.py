import io

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import CatalogVersion, DestinationProduct, Feedback, MasterProduct, Match, Upload
from app.services import standalone_matching
from app.services.search.index_manager import CatalogSearchIndex
from app.services.search.loader import load_master_records


@pytest.fixture
def db_session():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture(autouse=True)
def _isolated_qdrant_storage(tmp_path, monkeypatch):
    from app.config import settings
    from app.services.search import vector_search as vector_search_module

    monkeypatch.setattr(settings, "qdrant_local_path", str(tmp_path / "qdrant"))
    monkeypatch.setattr(settings, "qdrant_url", None)
    monkeypatch.setattr(settings, "embedding_provider", "tfidf")
    monkeypatch.setattr(settings, "high_confidence_threshold", 0.60)
    monkeypatch.setattr(settings, "low_confidence_threshold", 0.15)
    # Same reasoning as test_vector_search.py's own isolation fixture -
    # never let a test read or write the real default on-disk index build
    # cache (HANDOFF.md section 18).
    monkeypatch.setattr(vector_search_module, "_CACHE_DIR", tmp_path / ".index_cache")


def _seed(db_session):
    master_upload = Upload(filename="master.xlsx", upload_type="master", status="done")
    dest_upload = Upload(filename="dest.xlsx", upload_type="destination", status="done")
    db_session.add_all([master_upload, dest_upload])
    db_session.flush()

    master_products = [
        MasterProduct(
            upload_id=master_upload.id, source_row=2, external_id="001",
            product_name="Грелка резиновая", normalized_name="грелка резиновая",
            unit="шт.", price=4900, is_group_header=False, raw_data={},
        ),
        MasterProduct(
            upload_id=master_upload.id, source_row=3, external_id="002",
            product_name="Совершенно другой товар xyz", normalized_name="совершенно другой товар xyz",
            unit="шт.", price=1000, is_group_header=False, raw_data={},
        ),
    ]
    dest_products = [
        # exact normalized-name match
        DestinationProduct(
            upload_id=dest_upload.id, source_row=2, external_id=None,
            product_name="Грелка резиновая", normalized_name="грелка резиновая",
            quantity=2, price=4900, status="pending", raw_data={},
        ),
        # no realistic match at all
        DestinationProduct(
            upload_id=dest_upload.id, source_row=3, external_id=None,
            product_name="Абсолютно неродственный предмет qqq", normalized_name="абсолютно неродственный предмет qqq",
            quantity=1, price=500, status="pending", raw_data={},
        ),
    ]
    db_session.add_all(master_products + dest_products)
    db_session.commit()
    return master_upload, dest_upload, master_products, dest_products


def _build_index(db_session):
    records = load_master_records(db_session)
    index = CatalogSearchIndex()
    index.build(records)
    return index


def test_classify_exact_match_goes_to_auto_matched(db_session):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    bucket, record = standalone_matching._classify_destination_product(db_session, index, dest_products[0])

    assert bucket == "auto_matched"
    assert record["selected_catalog_id"] == master_products[0].id
    assert record["score"] == 0.99
    assert record["destination_id"] == dest_products[0].id


def test_build_job_result_produces_correct_stats_and_progress(db_session):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    job_id = standalone_matching.create_job(total=2)
    result = standalone_matching.build_job_result(db_session, index, dest_upload.id, job_id)

    assert result["stats"]["auto_matched"] + result["stats"]["needs_review"] + result["stats"]["no_match"] == 2
    assert result["stats"]["auto_matched"] >= 1  # the exact-match item

    job = standalone_matching.get_job(job_id)
    assert job.current == 2
    assert job.percent == 100


def test_job_registry_create_get_update():
    job_id = standalone_matching.create_job(total=10)
    job = standalone_matching.get_job(job_id)
    assert job is not None
    assert job.status == "running"
    assert job.total == 10

    standalone_matching._update_job(job_id, status="done", percent=100)
    assert standalone_matching.get_job(job_id).status == "done"


def test_get_job_returns_none_for_unknown_id():
    assert standalone_matching.get_job("nonexistent") is None


def test_save_results_creates_match_and_feedback_for_confirmed_decision(db_session):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)

    rows = [
        {
            "destination_id": dest_products[0].id,
            "destination_name": dest_products[0].product_name,
            "catalog_id": master_products[0].id,
            "catalog_name": master_products[0].product_name,
            "score": 0.99,
            "decision": "авто",
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    assert dest_products[0].status == "matched"
    match = db_session.query(Match).filter_by(destination_product_id=dest_products[0].id).one()
    assert match.master_product_id == master_products[0].id
    assert match.match_type == "auto_accepted"

    feedback = db_session.query(Feedback).filter_by(destination_product_id=dest_products[0].id).one()
    assert feedback.decision_type == "auto_accepted"
    assert feedback.selected_master_product_id == master_products[0].id


def test_save_results_marks_no_match_for_rejected_decision(db_session):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)

    rows = [
        {
            "destination_id": dest_products[1].id,
            "destination_name": dest_products[1].product_name,
            "catalog_id": None,
            "catalog_name": None,
            "score": 0.1,
            "decision": "без совпадения",
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    assert dest_products[1].status == "no_match"
    feedback = db_session.query(Feedback).filter_by(destination_product_id=dest_products[1].id).one()
    assert feedback.decision_type == "no_match"
    assert feedback.selected_master_product_id is None


def test_save_results_skips_unknown_destination_id_without_crashing(db_session):
    rows = [{"destination_id": "does-not-exist", "decision": "авто", "catalog_id": "also-fake"}]
    standalone_matching.save_results(db_session, rows)  # should not raise
    db_session.commit()


def test_save_results_does_not_persist_skipped_decision(db_session):
    """"Пропустить" (skip) during manual review is a deliberate "decide later":
    it must NOT create a Match or Feedback row and must NOT change the
    destination product's status, so the item can be revisited on a later run
    instead of being silently recorded as a no-match."""
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    target = dest_products[0]
    original_status = target.status

    rows = [
        {
            "destination_id": target.id,
            "destination_name": target.product_name,
            "catalog_id": None,
            "catalog_name": None,
            "score": 0.5,
            "decision": "пропущено",
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    assert target.status == original_status
    assert db_session.query(Match).filter_by(destination_product_id=target.id).count() == 0
    assert db_session.query(Feedback).filter_by(destination_product_id=target.id).count() == 0


def test_build_export_workbook_marks_skipped_rows():
    """A skipped row is exported as SKIPPED and flagged not-yet-reviewed, so a
    human can still find and finish those items in the output file.

    Column indices below match `_EXPORT_HEADERS` (destination name/desc/price,
    catalog code/name/desc/price, confidence %, type, reviewed, row id) - NOT
    the older English-header layout this test used to assert against. That
    layout was replaced when the export was localized to Russian (spec
    section 28); these two tests were never updated to match and were
    failing at HEAD before this change, unrelated to any catalog-versioning
    work.
    """
    rows = [
        {
            "destination_id": "d-skip",
            "destination_name": "Отложенный товар",
            "catalog_id": None,
            "catalog_name": None,
            "score": 0.5,
            "decision": "пропущено",
        }
    ]
    xlsx_bytes = standalone_matching.build_export_workbook(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert data_row[8] == "SKIPPED"  # "Тип" column
    assert data_row[9] == "No"       # "Проверено" column


class _FakeConfirmer:
    """Stub matching LLMAutoMatchConfirmer's BATCHED interface (HANDOFF.md
    section 10.2/10.4), for testing build_job_result's phase-2
    finalization logic in isolation from the real confirmer's own
    score-floor / chunking behavior (that's covered separately, against the
    real class, in test_reranking.py).
    """

    def __init__(self, results: list[bool]):
        self._results = results

    def confirm_batch(self, items):
        return self._results


def _fake_below_threshold_candidate(master_product, score: float):
    """A single CandidateWithProduct with a controlled final_score, low
    enough to fall below the (fixture-lowered) high_confidence_threshold,
    so tests can exercise the LLM-auto-match branch without depending on
    exactly what a real 2-row TF-IDF index happens to score.
    """
    from app.services.matching import CandidateWithProduct
    from app.services.search.types import ScoredCandidate

    return CandidateWithProduct(
        candidate=ScoredCandidate(master_product_id=master_product.id, final_score=score),
        master_product=master_product,
        explanation=[],
    )


def test_classify_defers_a_below_threshold_candidate_for_llm_confirmation(db_session, monkeypatch):
    """A candidate that's plausible (above needs_review territory) but not
    confident enough to clear the hybrid threshold on its own must be
    DEFERRED as a PendingLLMConfirmation, not decided inline - HANDOFF.md
    section 10.2/10.4's batching fix depends on every such row being
    collected across the whole run before any LLM call is made, rather
    than _classify_destination_product deciding (and calling the LLM) one
    row at a time the way it used to.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    import app.services.matching as matching_module

    fake_candidate = _fake_below_threshold_candidate(master_products[1], score=0.40)
    monkeypatch.setattr(matching_module, "find_exact_match", lambda *a, **k: None)
    monkeypatch.setattr(matching_module, "get_top_candidates", lambda *a, **k: [fake_candidate])

    result = standalone_matching._classify_destination_product(db_session, index, dest_products[1])

    assert isinstance(result, standalone_matching.PendingLLMConfirmation)
    assert result.top_score == 0.40
    assert result.best_master_product_id == master_products[1].id
    assert result.candidate_text == master_products[1].product_name


def test_build_job_result_llm_confirms_a_below_threshold_candidate_as_auto_matched(db_session, monkeypatch):
    """The case this feature exists for: a candidate that's plausible
    (above needs_review territory) but not confident enough to clear the
    hybrid threshold on its own - here a (fake, already-batched) LLM
    confirmation pushes it into auto_matched, tagged so it stays traceable
    as an LLM decision. Exercises build_job_result's phase-2 finalization;
    the real LLMAutoMatchConfirmer.confirm_batch is covered separately in
    test_reranking.py.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    pending = standalone_matching.PendingLLMConfirmation(
        dp=dest_products[1],
        candidates_json=[],
        query_text=dest_products[1].product_name,
        candidate_text=master_products[1].product_name,
        top_score=0.40,
        best_master_product_id=master_products[1].id,
    )
    monkeypatch.setattr(standalone_matching, "_classify_destination_product", lambda *a, **k: pending)
    monkeypatch.setattr(standalone_matching, "_build_auto_match_confirmer", lambda: _FakeConfirmer(results=[True, True]))

    job_id = standalone_matching.create_job(total=2)
    result = standalone_matching.build_job_result(db_session, index, dest_upload.id, job_id)

    assert result["stats"]["auto_matched"] == 2
    record = result["auto_matched"][0]
    assert record["auto_match_source"] == "llm_confirmed"
    assert record["selected_catalog_id"] == master_products[1].id
    assert record["score"] == 0.40


def test_build_job_result_llm_decline_falls_through_to_needs_review(db_session, monkeypatch):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    pending = standalone_matching.PendingLLMConfirmation(
        dp=dest_products[1],
        candidates_json=[],
        query_text=dest_products[1].product_name,
        candidate_text=master_products[1].product_name,
        top_score=0.40,
        best_master_product_id=master_products[1].id,
    )
    monkeypatch.setattr(standalone_matching, "_classify_destination_product", lambda *a, **k: pending)
    monkeypatch.setattr(standalone_matching, "_build_auto_match_confirmer", lambda: _FakeConfirmer(results=[False, False]))

    job_id = standalone_matching.create_job(total=2)
    result = standalone_matching.build_job_result(db_session, index, dest_upload.id, job_id)

    assert result["stats"]["needs_review"] == 2
    record = result["needs_review"][0]
    assert "auto_match_source" not in record


def test_build_job_result_batches_llm_confirmations_instead_of_one_call_per_row(db_session, monkeypatch):
    """The actual point of HANDOFF.md section 10.2/10.4: N eligible rows
    must produce ONE call to confirm_batch, not N calls to confirm - this
    is what a real 20-request/day free quota needs to survive a run with
    more than ~20 ambiguous rows.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    call_log: list[int] = []

    class _CountingConfirmer:
        def confirm_batch(self, items):
            call_log.append(len(items))
            return [True] * len(items)

    pending = standalone_matching.PendingLLMConfirmation(
        dp=dest_products[1],
        candidates_json=[],
        query_text=dest_products[1].product_name,
        candidate_text=master_products[1].product_name,
        top_score=0.40,
        best_master_product_id=master_products[1].id,
    )
    monkeypatch.setattr(standalone_matching, "_classify_destination_product", lambda *a, **k: pending)
    monkeypatch.setattr(standalone_matching, "_build_auto_match_confirmer", lambda: _CountingConfirmer())

    job_id = standalone_matching.create_job(total=2)
    standalone_matching.build_job_result(db_session, index, dest_upload.id, job_id)

    assert call_log == [2]  # one call covering both rows, not two calls of one each


def test_classify_exact_match_reports_its_own_auto_match_source(db_session):
    """Exact matches must also carry auto_match_source, distinctly from
    hybrid_threshold/llm_confirmed, so save_results can tell all three
    apart (see _AUTO_MATCH_SOURCE_TO_METHOD).
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    bucket, record = standalone_matching._classify_destination_product(db_session, index, dest_products[0])

    assert bucket == "auto_matched"
    assert record["auto_match_source"] == "exact_match"


def test_save_results_stamps_llm_auto_matched_method(db_session):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)

    rows = [
        {
            "destination_id": dest_products[1].id,
            "destination_name": dest_products[1].product_name,
            "catalog_id": master_products[1].id,
            "catalog_name": master_products[1].product_name,
            "score": 0.76,
            "decision": "авто",
            "auto_match_source": "llm_confirmed",
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    match = db_session.query(Match).filter_by(destination_product_id=dest_products[1].id).one()
    assert match.method == "llm_auto_matched"
    assert match.match_type == "auto_accepted"  # unchanged - only `method` distinguishes the source


def test_save_results_stamps_hybrid_threshold_method(db_session):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)

    rows = [
        {
            "destination_id": dest_products[1].id,
            "catalog_id": master_products[1].id,
            "score": 0.7,
            "decision": "авто",
            "auto_match_source": "hybrid_threshold",
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    match = db_session.query(Match).filter_by(destination_product_id=dest_products[1].id).one()
    assert match.method == "auto_accepted_hybrid"


def test_save_results_defaults_to_exact_match_method_when_source_missing(db_session):
    """Backward compatibility: a caller that predates this feature (an
    older frontend build, or a direct API call) sends no auto_match_source
    at all - must not crash, and must fall back to the original ("авто" ->
    "exact_match") behavior rather than some new default.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)

    rows = [
        {
            "destination_id": dest_products[0].id,
            "catalog_id": master_products[0].id,
            "score": 0.99,
            "decision": "авто",
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    match = db_session.query(Match).filter_by(destination_product_id=dest_products[0].id).one()
    assert match.method == "exact_match"


def test_build_export_workbook_marks_llm_confirmed_rows_distinctly():
    rows = [
        {
            "destination_id": "d1",
            "destination_name": "Дозатор для житкого мыла",
            "catalog_id": "m1",
            "catalog_name": "Дозатор жидкого мыла",
            "score": 0.762,
            "decision": "авто",
            "auto_match_source": "llm_confirmed",
        },
        {
            "destination_id": "d2",
            "destination_name": "Стол",
            "catalog_id": "m2",
            "catalog_name": "Стол",
            "score": 0.90,
            "decision": "авто",
            "auto_match_source": "hybrid_threshold",
        },
    ]
    xlsx_bytes = standalone_matching.build_export_workbook(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    row1 = [cell.value for cell in ws[2]]
    row2 = [cell.value for cell in ws[3]]
    assert row1[8] == "AUTO_MATCH_AI"   # Тип
    assert row2[8] == "AUTO_MATCH"      # plain hybrid-threshold auto-match stays as before


def test_load_master_records_scoped_to_upload_id(db_session):
    """Real bug found in production: the standalone wizard creates a brand
    new master upload every run but never deletes old ones, and the search
    index used to be rebuilt from *every* MasterProduct row ever ingested
    (see loader.py before the fix). If an earlier run (or a mistaken file
    drop) had ingested a different catalog - e.g. the destination file
    itself - its rows stayed in the table forever and silently leaked into
    every later run's matches. This locks in the fix: passing `upload_id`
    to `load_master_records` returns only that upload's rows.
    """
    old_upload = Upload(filename="old_wrong_catalog.xlsx", upload_type="master", status="done")
    new_upload = Upload(filename="real_catalog.xlsx", upload_type="master", status="done")
    db_session.add_all([old_upload, new_upload])
    db_session.flush()

    poison = MasterProduct(
        upload_id=old_upload.id, source_row=2, external_id="P1",
        product_name="Товар из старой загрузки", normalized_name="товар из старой загрузки",
        unit="шт.", price=1, is_group_header=False, raw_data={},
    )
    real = MasterProduct(
        upload_id=new_upload.id, source_row=2, external_id="R1",
        product_name="Настоящий товар каталога", normalized_name="настоящий товар каталога",
        unit="шт.", price=2, is_group_header=False, raw_data={},
    )
    db_session.add_all([poison, real])
    db_session.commit()

    unscoped = load_master_records(db_session)
    assert {r.id for r in unscoped} == {poison.id, real.id}

    scoped = load_master_records(db_session, upload_id=new_upload.id)
    assert {r.id for r in scoped} == {real.id}


@pytest.fixture
def shared_engine_session(tmp_path):
    """A StaticPool-backed in-memory engine so multiple SessionLocal()
    calls (as run_matching_job makes internally) all see the same seeded
    data - a plain ":memory:" engine without StaticPool would hand back a
    fresh, empty database on every new connection.
    """
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool

    engine = create_engine(
        "sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    SessionFactory = sessionmaker(bind=engine)
    session = SessionFactory()
    yield session, SessionFactory
    session.close()


def test_run_matching_job_does_not_match_against_a_different_earlier_upload(
    shared_engine_session, monkeypatch, tmp_path
):
    """End-to-end version of the scoping fix: seed an old, unrelated master
    upload whose product name exactly equals a destination product's name
    (simulating the real production incident - Детсад having been ingested
    as a "catalog" file in an earlier run), plus a genuine, freshly-ingested
    master catalog that does NOT contain that name. The destination product
    must NOT come back as an exact/auto match, since the only catalog
    actually uploaded *this run* has no such product.
    """
    session, SessionFactory = shared_engine_session
    import app.services.standalone_matching as sm_module

    monkeypatch.setattr(sm_module, "SessionLocal", SessionFactory)

    old_upload = Upload(filename="old_wrong_catalog.xlsx", upload_type="master", status="done")
    session.add(old_upload)
    session.flush()
    session.add(
        MasterProduct(
            upload_id=old_upload.id, source_row=2, external_id="P1",
            product_name="Аппарат для гальванизации", normalized_name="аппарат для гальванизации",
            unit="шт.", price=1, is_group_header=False, raw_data={},
        )
    )
    session.commit()

    # Build a tiny real master + destination workbook pair for ingest_master /
    # ingest_destination to read from disk.
    master_wb = openpyxl.Workbook()
    ws = master_wb.active
    ws.append(["Код", "Наименование", "Единица измерения", "Цена"])
    ws.append(["M1", "Совершенно другой настоящий товар", "шт.", 500])
    master_path = tmp_path / "real_catalog.xlsx"
    master_wb.save(master_path)

    dest_wb = openpyxl.Workbook()
    ws = dest_wb.active
    ws.append(["Наименование товара", "Описание", "Цена", "Кол-во"])
    ws.append(["Аппарат для гальванизации", "любое описание", 100, 1])
    dest_path = tmp_path / "destination.xlsx"
    dest_wb.save(dest_path)

    job_id = sm_module.create_job(total=1)
    sm_module.run_matching_job(job_id, str(master_path), "real_catalog.xlsx", str(dest_path), "destination.xlsx")

    job = sm_module.get_job(job_id)
    assert job.status == "done", job.error
    all_records = [
        r
        for bucket in ("auto_matched", "needs_review", "no_match")
        for r in job.result[bucket]
    ]
    matched = next(r for r in all_records if r["destination_name"] == "Аппарат для гальванизации")

    # Must NOT have auto-matched against the old, unrelated upload's product -
    # the real catalog uploaded this run has no such item, so this should
    # land in needs_review or no_match, not a 0.99-confidence exact match.
    assert matched.get("selected_catalog_id") != "P1"
    assert matched.get("score") != 0.99


def test_build_export_workbook_has_expected_headers_and_rows(tmp_path):
    """Locks in the actual (Russian, spec section 28) export layout - see
    _EXPORT_HEADERS. This test used to assert an older English-header
    layout that no longer matches build_export_workbook and was failing at
    HEAD before this change, unrelated to any catalog-versioning work.
    """
    rows = [
        {
            "destination_id": "d1",
            "destination_name": "Грелка резиновая",
            "catalog_id": "m1",
            "catalog_name": "Грелка резиновая",
            "score": 0.99,
            "decision": "авто",
        },
        {
            "destination_id": "d2",
            "destination_name": "Что-то без пары",
            "catalog_id": None,
            "catalog_name": None,
            "score": 0.1,
            "decision": "без совпадения",
        },
    ]
    xlsx_bytes = standalone_matching.build_export_workbook(rows)

    out_path = tmp_path / "export.xlsx"
    out_path.write_bytes(xlsx_bytes)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert header == standalone_matching._EXPORT_HEADERS

    row1 = [cell.value for cell in ws[2]]
    assert row1[0] == "Грелка резиновая"  # Наименование (заявка)
    assert row1[3] == "m1"           # Код (каталог) - falls back to catalog_id when no catalog_code
    assert row1[7] == "99%"          # Совпадение %
    assert row1[8] == "AUTO_MATCH"   # Тип
    assert row1[9] == "No"           # Проверено - "авто" decisions are not reviewed-by-human
    assert row1[10] == "d1"          # ID строки

    row2 = [cell.value for cell in ws[3]]
    # openpyxl round-trips a written "" as a blank cell (None) on reload,
    # not as an empty string - "falsy" covers both without depending on
    # that library detail.
    assert not row2[3]                # no catalog_id or catalog_code at all
    assert row2[7] == "10%"
    assert row2[8] == "NO_MATCH"
    assert row2[9] == "Yes"


# --- "No resume" fix (NEXT_STEPS.md item 6) --------------------------------
#
# Part A: every wizard decision persists the instant it's made (POST
# /decisions -> save_results with already_persisted checked at final /save
# time so nothing is written twice). Part B: reopening an existing
# destination upload skips reclassifying already-decided rows.


def test_save_results_skips_rows_already_persisted(db_session):
    """Part A's whole point: a decision saved instantly via /decisions must
    not be written AGAIN when the final /save call includes it a second
    time - otherwise every session would double (or more) its own Match/
    Feedback rows.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    rows = [
        {
            "destination_id": dest_products[0].id,
            "catalog_id": master_products[0].id,
            "score": 0.95,
            "decision": "вручную",
            "already_persisted": True,
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    assert dest_products[0].status == "pending"  # untouched - the row was never actually written
    assert db_session.query(Match).filter_by(destination_product_id=dest_products[0].id).count() == 0
    assert db_session.query(Feedback).filter_by(destination_product_id=dest_products[0].id).count() == 0


def test_save_results_still_persists_rows_not_flagged_already_persisted(db_session):
    """The flag must be opt-in, not opt-out - every existing caller (and
    every row a reviewer decides right at final-save time, e.g. an
    auto-match override) has no `already_persisted` key at all and must
    keep writing exactly as before.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    rows = [
        {
            "destination_id": dest_products[0].id,
            "catalog_id": master_products[0].id,
            "score": 0.95,
            "decision": "вручную",
        }
    ]
    standalone_matching.save_results(db_session, rows)
    db_session.commit()

    assert dest_products[0].status == "matched"
    assert db_session.query(Match).filter_by(destination_product_id=dest_products[0].id).count() == 1


def test_resume_record_for_manually_matched_row(db_session):
    """_resume_record_for_decided must report the REAL decision (a Match
    with a real master_product_id, match_type="user_selected") rather than
    reclassifying - this is the core correctness property Part B depends
    on: a previously-confirmed choice must never silently change.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    dp = dest_products[1]  # the one with no realistic auto-match
    dp.status = "matched"
    db_session.add(
        Match(
            destination_product_id=dp.id,
            master_product_id=master_products[1].id,
            confidence=0.72,
            match_type="user_selected",
            method="standalone_wizard",
            is_confirmed=True,
        )
    )
    db_session.commit()

    bucket, record = standalone_matching._resume_record_for_decided(db_session, dp)

    assert bucket == "auto_matched"
    assert record["already_decided"] is True
    assert record["resumed_decision"] == "вручную"
    assert record["selected_catalog_id"] == master_products[1].id
    assert record["score"] == 0.72


def test_resume_record_for_rejected_row(db_session):
    """A row explicitly marked "no_match" (an earlier "Не подходит") with no
    Match at all must resume into the no_match bucket, not get reclassified
    into a fresh candidate search.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    dp = dest_products[0]
    dp.status = "no_match"
    db_session.commit()

    bucket, record = standalone_matching._resume_record_for_decided(db_session, dp)

    assert bucket == "no_match"
    assert record["already_decided"] is True
    assert record["resumed_decision"] == "без совпадения"
    assert record["candidates"] == []


def test_build_job_result_does_not_reclassify_already_decided_rows(db_session):
    """The actual integration point: build_job_result must route a
    non-pending row through _resume_record_for_decided instead of
    _classify_destination_product, and a genuinely pending row through the
    normal classification path unchanged.
    """
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    index = _build_index(db_session)

    decided = dest_products[0]
    decided.status = "matched"
    db_session.add(
        Match(
            destination_product_id=decided.id,
            master_product_id=master_products[1].id,  # deliberately NOT what fresh classification would pick
            confidence=0.5,
            match_type="user_selected",
            method="standalone_wizard",
            is_confirmed=True,
        )
    )
    db_session.commit()

    job_id = standalone_matching.create_job(total=2)
    result = standalone_matching.build_job_result(db_session, index, dest_upload.id, job_id)

    all_records = [r for bucket in ("auto_matched", "needs_review", "no_match") for r in result[bucket]]
    resumed = next(r for r in all_records if r["destination_id"] == decided.id)

    # Must reflect the STORED decision (master_products[1]), not whatever
    # fresh classification would have picked for "Грелка резиновая" (which
    # has an exact-match candidate at master_products[0] - a real
    # disagreement, deliberately chosen for this test so a bug that ignores
    # already_decided would be caught, not accidentally pass).
    assert resumed.get("already_decided") is True
    assert resumed.get("selected_catalog_id") == master_products[1].id

    still_pending = next(r for r in all_records if r["destination_id"] == dest_products[1].id)
    assert not still_pending.get("already_decided")


def test_list_resumable_destinations_reports_correct_counts(db_session):
    master_upload, dest_upload, master_products, dest_products = _seed(db_session)
    dest_products[0].status = "matched"
    db_session.commit()

    resumable = standalone_matching.list_resumable_destinations(db_session)

    entry = next(d for d in resumable if d["upload_id"] == dest_upload.id)
    assert entry["total"] == 2
    assert entry["decided"] == 1
    assert entry["pending"] == 1


def test_list_resumable_destinations_excludes_empty_uploads(db_session):
    empty_upload = Upload(filename="empty.xlsx", upload_type="destination", status="done")
    db_session.add(empty_upload)
    db_session.commit()

    resumable = standalone_matching.list_resumable_destinations(db_session)

    assert all(d["upload_id"] != empty_upload.id for d in resumable)


def test_run_matching_job_resume_mode_does_not_reingest_destination(
    shared_engine_session, monkeypatch, tmp_path
):
    """Part B's other half: passing destination_upload_id must reuse the
    existing DestinationProduct rows (and their ids) rather than creating a
    second, parallel set via a fresh ingest_destination() call - otherwise
    "resume" would just quietly duplicate the whole destination file.
    """
    session, SessionFactory = shared_engine_session
    import app.services.standalone_matching as sm_module

    monkeypatch.setattr(sm_module, "SessionLocal", SessionFactory)

    master_wb = openpyxl.Workbook()
    ws = master_wb.active
    ws.append(["Код", "Наименование", "Единица измерения", "Цена"])
    ws.append(["M1", "Совершенно другой настоящий товар", "шт.", 500])
    master_path = tmp_path / "real_catalog.xlsx"
    master_wb.save(master_path)

    dest_wb = openpyxl.Workbook()
    ws = dest_wb.active
    ws.append(["Наименование товара", "Описание", "Цена", "Кол-во"])
    ws.append(["Аппарат для гальванизации", "любое описание", 100, 1])
    dest_path = tmp_path / "destination.xlsx"
    dest_wb.save(dest_path)

    # First run: fresh ingest of both files, creating the real DestinationProduct row.
    job_id_1 = sm_module.create_job(total=1)
    sm_module.run_matching_job(job_id_1, str(master_path), "real_catalog.xlsx", str(dest_path), "destination.xlsx")
    job_1 = sm_module.get_job(job_id_1)
    assert job_1.status == "done", job_1.error

    dest_upload_id = session.query(Upload).filter_by(upload_type="destination").one().id
    original_dp_id = session.query(DestinationProduct).filter_by(upload_id=dest_upload_id).one().id
    catalog_version_id = job_1.result["catalog_version_id"]

    # Second run: RESUME the same destination upload, reusing the catalog
    # version instead of re-uploading either file.
    job_id_2 = sm_module.create_job(total=1)
    sm_module.run_matching_job(
        job_id_2, None, None, None, None,
        catalog_version_id=catalog_version_id,
        destination_upload_id=dest_upload_id,
    )
    job_2 = sm_module.get_job(job_id_2)
    assert job_2.status == "done", job_2.error

    # Still exactly one DestinationProduct row for this upload - resuming
    # must not have re-ingested a second copy of the file.
    dps = session.query(DestinationProduct).filter_by(upload_id=dest_upload_id).all()
    assert len(dps) == 1
    assert dps[0].id == original_dp_id


# --- Manual catalog search in the wizard (NEXT_STEPS.md item 6) ------------


def _seed_two_catalogs(db_session):
    """Two separate catalogs (like April vs. an unrelated one) with
    DISJOINT products - the exact setup needed to prove manual_search()
    only ever searches the ONE catalog a run is actually using, not
    whatever else happens to be in the database.
    """
    upload_a = Upload(filename="catalog_a.xlsx", upload_type="master", status="done")
    upload_b = Upload(filename="catalog_b.xlsx", upload_type="master", status="done")
    db_session.add_all([upload_a, upload_b])
    db_session.flush()

    product_a = MasterProduct(
        upload_id=upload_a.id, source_row=2, external_id="A1",
        product_name="Стол офисный регулируемый", normalized_name="стол офисный регулируемый",
        unit="шт.", price=15000, is_group_header=False, raw_data={},
    )
    product_b = MasterProduct(
        upload_id=upload_b.id, source_row=2, external_id="B1",
        product_name="Шкаф архивный металлический", normalized_name="шкаф архивный металлический",
        unit="шт.", price=25000, is_group_header=False, raw_data={},
    )
    db_session.add_all([product_a, product_b])
    db_session.flush()

    version_a = CatalogVersion(name="Catalog A", source_upload_id=upload_a.id, is_active=True, product_count=1)
    version_b = CatalogVersion(name="Catalog B", source_upload_id=upload_b.id, is_active=True, product_count=1)
    db_session.add_all([version_a, version_b])
    db_session.commit()
    return version_a, version_b, product_a, product_b


def test_manual_search_returns_results_for_the_target_catalog(db_session):
    version_a, _version_b, product_a, _product_b = _seed_two_catalogs(db_session)

    results = standalone_matching.manual_search(db_session, version_a, "стол офисный", top_k=5)

    assert len(results) >= 1
    assert results[0]["catalog_id"] == product_a.id
    assert results[0]["product_name"] == "Стол офисный регулируемый"
    assert results[0]["external_id"] == "A1"


def test_qdrant_collection_name_is_distinct_per_upload():
    """The exact invariant NEXT_STEPS.md section 6a's fix depends on -
    see standalone_matching._qdrant_collection_name's own docstring for
    the real bug this exists to prevent (two catalogs used to share one
    hardcoded Qdrant collection name; building a second one silently wiped
    an older, still-cached catalog's real vector data).
    """
    assert standalone_matching._qdrant_collection_name("upload-a") != standalone_matching._qdrant_collection_name("upload-b")
    assert standalone_matching._qdrant_collection_name("upload-a") == standalone_matching._qdrant_collection_name("upload-a")


def test_manual_search_never_returns_results_from_a_different_catalog_version(db_session, monkeypatch, tmp_path):
    """The actual reason this function exists instead of reusing
    app/api/matching.py's global-index manual search - see
    standalone_matching.manual_search's own docstring.

    Separate, mechanical reason this test needs its own local Qdrant path
    for version_b's build, UNRELATED to the collection-name fix above:
    embedded/local-mode Qdrant exclusively locks its storage DIRECTORY (not
    per-collection), so two simultaneously-live clients (version_a's stays
    cached in _version_index_cache, version_b's is freshly opened) crash
    with "already accessed by another instance" if they share a path - the
    default the test isolation fixture gives every test. This has nothing
    to do with which COLLECTION each targets (the actual fix under test
    here) and everything to do with local mode's own file lock; the real
    deployment uses a real Qdrant SERVER, which has no such restriction and
    happily serves concurrent connections - local mode's own error message
    says as much ("If you require concurrent access, use Qdrant server
    instead"). See test_qdrant_collection_name_is_distinct_per_upload
    above for the narrower, direct, Qdrant-free check of the actual fix
    this test can't fully exercise itself under local mode's constraints.
    """
    version_a, version_b, product_a, product_b = _seed_two_catalogs(db_session)

    results_a = standalone_matching.manual_search(db_session, version_a, "шкаф архивный", top_k=5)
    assert all(r["catalog_id"] != product_b.id for r in results_a)

    from app.config import settings
    monkeypatch.setattr(settings, "qdrant_local_path", str(tmp_path / "qdrant_b"))
    results_b = standalone_matching.manual_search(db_session, version_b, "стол офисный", top_k=5)
    assert all(r["catalog_id"] != product_a.id for r in results_b)


def test_manual_search_rebuilds_the_index_on_a_cache_miss(db_session):
    """No prior run has cached this version's index (e.g. right after a
    server restart) - manual_search must still work, building and caching
    it on the fly rather than raising.
    """
    version_a, _version_b, product_a, _product_b = _seed_two_catalogs(db_session)

    from app.services.search.index_manager import get_cached_index_for_version
    assert get_cached_index_for_version(version_a.id) is None  # precondition: genuinely uncached

    results = standalone_matching.manual_search(db_session, version_a, "стол", top_k=5)
    assert any(r["catalog_id"] == product_a.id for r in results)
    assert get_cached_index_for_version(version_a.id) is not None  # now warmed for next time


def test_manual_search_uses_the_already_cached_index_when_present(db_session):
    """A run already in progress (run_matching_job already built and
    cached this version's index) must have its manual searches hit that
    SAME index/data - not silently rebuild from a possibly-stale DB read.
    """
    version_a, _version_b, product_a, _product_b = _seed_two_catalogs(db_session)
    index = _build_index(db_session)  # scoped to ALL master products by default - fine for this check
    from app.services.search.index_manager import cache_index_for_version
    cache_index_for_version(version_a.id, index)

    results = standalone_matching.manual_search(db_session, version_a, "стол офисный", top_k=5)
    assert any(r["catalog_id"] == product_a.id for r in results)
