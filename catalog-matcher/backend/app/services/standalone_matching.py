"""Backend adapter for the standalone matching wizard tab
(`matching-standalone.html`, served from frontend/public and embedded as
its own tab in the Next.js app).

That HTML/JS file is a self-contained tool with its own simple contract:
upload a destination file + a catalog file together, run one matching
pass, review only the uncertain ones in a one-question-at-a-time wizard,
then save. It expects a specific JSON shape (see `_classify_destination`
and `build_job_result`) and a background job it can poll for progress,
neither of which anything else in this project needed before - so this
module exists to translate that contract onto the same real
ingestion/search/matching services every other tab already uses, rather
than building a second, parallel matching engine.

Every run here creates fresh `Upload`/`MasterProduct`/`DestinationProduct`
rows (the standalone tool has no "pick a previous upload" concept - it's
a one-shot "upload both, get an answer" tool), and nothing is written to
`matches`/`feedback` until the user actually clicks Save, matching the
tool's own preview-then-commit UX.
"""

from __future__ import annotations

import io
import threading
import uuid
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import settings
from app.database import SessionLocal
from app.models import CatalogVersion, DestinationProduct, Feedback, MasterProduct, Match, Upload
from app.services import matching
from app.services.ingestion import IngestionOptions, ingest_destination, ingest_master
from app.services.search.index_manager import (
    CatalogSearchIndex,
    cache_index_for_version,
    get_cached_index_for_version,
)
from app.services.search.loader import load_master_records
from app.services.search.reranking import LLMAutoMatchConfirmer, _build_llm_client

# --- In-memory job registry -------------------------------------------------
#
# A single dev-server process is all this project targets (see
# CatalogSearchIndex's own singleton for the same trade-off). Jobs are lost
# on restart; that's fine for a one-shot "run this now" tool where nothing
# is ever left half-done across a server restart, unlike the main review
# workflow's persisted `destination_products.status`.


@dataclass
class JobState:
    status: str = "running"  # "running" | "done" | "error"
    percent: int = 0
    current: int = 0
    total: int = 0
    result: dict | None = None
    error: str | None = None


_jobs: dict[str, JobState] = {}
_jobs_lock = threading.Lock()


def create_job(total: int) -> str:
    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = JobState(total=total)
    return job_id


def get_job(job_id: str) -> JobState | None:
    with _jobs_lock:
        return _jobs.get(job_id)


def _update_job(job_id: str, **kwargs) -> None:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job is None:
            return
        for key, value in kwargs.items():
            setattr(job, key, value)


# --- Classification (auto_matched / needs_review / no_match) ---------------


def _destination_json(dp: DestinationProduct) -> dict[str, Any]:
    """The requested item's own fields, shared by every bucket.

    Factored out because all four branches of
    `_classify_destination_product` return the same destination fields;
    duplicating them meant every new field had to be added in four places
    and was easy to miss in one.

    `source_row` is included so the UI can show "стр. 42" when the
    destination file has no code column of its own - which is the normal
    case for these request lists. Without it the only identifier available
    to the frontend is the internal UUID, which is meaningless to a
    reviewer trying to find the row in Excel.
    """
    price = dp.price
    return {
        "destination_id": dp.id,
        "destination_name": dp.product_name,
        "destination_description": dp.description,
        "destination_external_id": dp.external_id,
        "destination_source_row": dp.source_row,
        "destination_price": float(price) if price is not None else None,
        "destination_quantity": float(dp.quantity) if dp.quantity is not None else None,
    }


def _candidate_json(cwp: matching.CandidateWithProduct) -> dict[str, Any]:
    """Payload for one candidate shown in the wizard.

    `description` (plus code/unit/price) is included so the reviewer can
    verify a match without opening the source spreadsheet. Catalog names
    are frequently generic - "Стол рабочий на 4 детей", "Шкаф стеллаж
    открытый" - and the detail that separates two near-identical entries
    (dimensions, material, class) exists only in the description column.
    """
    price = cwp.master_product.price
    return {
        "catalog_id": cwp.master_product.id,
        "product_name": cwp.master_product.product_name,
        "description": cwp.master_product.description,
        "external_id": cwp.master_product.external_id,
        "unit": cwp.master_product.unit,
        "price": float(price) if price is not None else None,
        "score": cwp.candidate.final_score,
        "coverage": cwp.candidate.lexical_overlap_score,
    }


def manual_search(db: Session, catalog_version: CatalogVersion, query: str, top_k: int = 10) -> list[dict[str, Any]]:
    """Manual catalog search for the wizard (NEXT_STEPS.md item 6, second
    bullet: "no manual catalog search in the wizard"). If none of the 3
    auto-suggested candidates fit, this is what lets a reviewer search the
    WHOLE catalog and pick a different item, instead of the only previous
    option, "Не подходит".

    Deliberately does NOT reuse app/api/matching.py's own `/manual-search`
    endpoint (the old Review tab's) - that one queries
    `index_manager.get_index()`, the catalog-agnostic GLOBAL singleton
    index, which is not necessarily the catalog this wizard run is even
    using. This scopes to `catalog_version`'s own per-version cached index
    (same cache `run_matching_job` reads/writes - `get_cached_index_for_version`/
    `cache_index_for_version`), so a search always returns results from
    the SAME catalog the reviewer is actually working against - rebuilding
    once, exactly like run_matching_job's own cache-miss path, if nothing
    has warmed it yet (e.g. right after a server restart).

    Uses `index.search()` (raw Phase 2 retrieval), not `search_reranked()` -
    a manual search is the reviewer explicitly overriding the algorithm's
    own top-3 guess, so it should show the broadest reasonable candidate
    set rather than the same reranking logic that already decided these
    weren't the top 3 in the first place.
    """
    index = get_cached_index_for_version(catalog_version.id)
    if index is None:
        records = load_master_records(db, upload_id=catalog_version.source_upload_id)
        index = CatalogSearchIndex()
        index.build(records, collection_name=_qdrant_collection_name(catalog_version.source_upload_id))
        cache_index_for_version(catalog_version.id, index)

    results = []
    for c in index.search(query, top_k=top_k):
        master_product = db.get(MasterProduct, c.master_product_id)
        if master_product is None:
            continue  # stale point in the index pointing at a since-deleted row - skip, don't crash
        cwp = matching.CandidateWithProduct(
            candidate=c, master_product=master_product, explanation=matching.build_explanation(c)
        )
        results.append(_candidate_json(cwp))
    return results


def _qdrant_collection_name(upload_id: str) -> str:
    """Distinct Qdrant collection per catalog (NEXT_STEPS.md section 6a) -
    keyed by the catalog's master Upload id, not its CatalogVersion id,
    because the fresh-upload branch of `run_matching_job` builds the index
    BEFORE the CatalogVersion row exists (`_create_catalog_version` needs
    the build's own stats.indexed_records) - the upload id is the one
    identifier that's always available at build time either way (a fresh
    upload or a `catalog_version.source_upload_id` when reusing one).

    See CatalogSearchIndex.build()'s own docstring for the full "why" -
    every catalog used to share ONE hardcoded collection name, and a
    second catalog's build would silently wipe an older, still-cached
    catalog's real Qdrant data.
    """
    return f"catalog_{upload_id}"


@dataclass
class PendingLLMConfirmation:
    """A destination row whose classification isn't final yet because it
    depends on an LLM confirmation call (HANDOFF.md section 10.2/10.4).

    `_classify_destination_product` used to make that call inline, one row
    at a time - which is exactly what exhausted a real 20-request/day free
    Gemini quota in under a minute (every remaining row then silently
    fell back to manual review via a 429, fail-safe but invisible unless
    you went looking at the logs). Now it returns this instead, and
    `build_job_result` collects every pending row across the WHOLE run
    before flushing them all through `LLMAutoMatchConfirmer.confirm_batch`
    in groups of `config.py`'s `llm_batch_size` - one API call covers many
    rows instead of one.

    Carries everything needed both to ask the LLM (`query_text`,
    `candidate_text`, `top_score`) and to finalize the row afterwards either
    way - confirmed (-> "auto_matched", source "llm_confirmed") or not
    (-> "needs_review" or "no_match", same score-based split
    `_classify_destination_product` always used).
    """

    dp: DestinationProduct
    candidates_json: list[dict[str, Any]]
    query_text: str
    candidate_text: str
    top_score: float
    best_master_product_id: str


def _resume_record_for_decided(db: Session, dp: DestinationProduct) -> tuple[str, dict[str, Any]]:
    """For a DestinationProduct whose status is no longer "pending" - i.e. a
    real decision already exists, either from earlier in THIS run (Part A
    of the "no resume" fix already wrote it via /decisions the instant it
    was made) or from a genuinely earlier session (Part B - resuming a
    previous destination upload) - report THAT decision directly instead of
    reclassifying the row fresh.

    This matters for a real reason, not just efficiency: a fresh
    reclassification could legitimately disagree with a decision a human
    already confirmed (an index rebuild, a recalibrated threshold, a
    different embedding provider - see HANDOFF.md section 16's LaBSE
    switch, which changes every score). A previously-confirmed choice must
    never silently change or get re-asked - re-deciding it is a separate,
    explicit action (the same "destructive-looking things must be
    deliberate" principle as MasterProduct.is_active elsewhere in this
    project), not a side effect of reopening the wizard.

    Bucketed as "auto_matched" (has a real master_product_id) or "no_match"
    (does not) purely so it's excluded from `needs_review`/the review
    queue and still shows up somewhere in the summary/detail screens - see
    `already_decided`/`resumed_decision` on the returned record, which the
    frontend reads to render a distinct badge and to skip re-persisting it
    at final save time (`save_results`'s `already_persisted` flag).
    """
    match = (
        db.query(Match)
        .filter(Match.destination_product_id == dp.id)
        .order_by(Match.created_at.desc())
        .first()
    )
    base = _destination_json(dp)
    base["already_decided"] = True

    if match is not None and match.master_product_id:
        master_product = db.get(MasterProduct, match.master_product_id)
        price = master_product.price if master_product else None
        confidence = float(match.confidence) if match.confidence is not None else None
        base["candidates"] = [
            {
                "catalog_id": match.master_product_id,
                "product_name": master_product.product_name if master_product else None,
                "description": master_product.description if master_product else None,
                "external_id": master_product.external_id if master_product else None,
                "unit": master_product.unit if master_product else None,
                "price": float(price) if price is not None else None,
                "score": confidence,
                "coverage": None,
            }
        ]
        base["selected_catalog_id"] = match.master_product_id
        base["score"] = confidence
        base["auto_match_source"] = None
        # match_type is "user_selected" (a real manual pick) or
        # "auto_accepted" (see _DECISION_TO_TYPE) - mapped back to the same
        # decision vocabulary buildSaveRows()/save_results already use.
        base["resumed_decision"] = "вручную" if match.match_type == "user_selected" else "авто"
        return "auto_matched", base

    base["candidates"] = []
    base["score"] = None
    base["resumed_decision"] = "без совпадения"
    return "no_match", base


def _classify_destination_product(
    db: Session,
    index: CatalogSearchIndex,
    dp: DestinationProduct,
    master_upload_id: str | None = None,
) -> tuple[str, dict[str, Any]] | PendingLLMConfirmation:
    """Returns either a decided (bucket, record_json) - bucket one of
    "auto_matched" (exact or hybrid-threshold only; never llm_confirmed,
    see below), "needs_review", "no_match" - or a `PendingLLMConfirmation`
    when the row still needs an LLM confirmation call to know its final
    bucket (HANDOFF.md section 10.2/10.4 - that call is no longer made
    here; see `PendingLLMConfirmation` and `build_job_result`).

    Reuses the exact same thresholds (HIGH_CONFIDENCE_THRESHOLD /
    LOW_CONFIDENCE_THRESHOLD) as the main review tab's Phase 5 auto-match,
    so this tool's behavior matches whatever the user already tuned there.

    `master_upload_id` scopes the exact-match check (see
    `matching.find_exact_match`) to just this run's freshly-ingested master
    upload - the same scoping already applied to `index` via
    `load_master_records(db, upload_id=...)`. Without this, exact-match
    still bypassed the index and queried MasterProduct directly across
    every upload ever ingested, so an old/unrelated master upload sitting
    in the table could produce a bogus 99% match even after the index-side
    fix.

    Every "auto_matched" record carries `auto_match_source` - "exact_match",
    "hybrid_threshold", or (once `build_job_result` finalizes a confirmed
    `PendingLLMConfirmation`) "llm_confirmed" - so `save_results` can stamp
    a distinct, traceable `Match.method` for each rather than collapsing
    all three into one indistinguishable label.
    """
    exact = matching.find_exact_match(db, dp, master_upload_id=master_upload_id)
    candidates = matching.get_top_candidates(db, index, dp, top_k=3)
    candidates_json = [_candidate_json(c) for c in candidates]

    if exact is not None:
        exact_price = exact.price
        return "auto_matched", {
            **_destination_json(dp),
            # Built by hand rather than via _candidate_json because an
            # exact match has no ScoredCandidate behind it (it bypasses
            # hybrid scoring entirely - see matching.find_exact_match), so
            # there are no sub-scores to report. Same keys, so the UI can
            # render it identically.
            "candidates": [
                {
                    "catalog_id": exact.id,
                    "product_name": exact.product_name,
                    "description": exact.description,
                    "external_id": exact.external_id,
                    "unit": exact.unit,
                    "price": float(exact_price) if exact_price is not None else None,
                    "score": 0.99,
                    "coverage": 1.0,
                }
            ]
            + [c for c in candidates_json if c["catalog_id"] != exact.id],
            "selected_catalog_id": exact.id,
            "score": 0.99,
            "auto_match_source": "exact_match",
        }

    top_score = candidates[0].candidate.final_score if candidates else 0.0

    if candidates and top_score >= settings.high_confidence_threshold:
        best = candidates[0]
        return "auto_matched", {
            **_destination_json(dp),
            "candidates": candidates_json,
            "selected_catalog_id": best.master_product.id,
            "score": top_score,
            "auto_match_source": "hybrid_threshold",
        }

    if candidates:
        # Deferred, not decided here - see PendingLLMConfirmation. Whether
        # this row ever actually reaches the LLM (vs. `confirm_batch`
        # skipping it below `medium_confidence_threshold`) and what bucket
        # it lands in if not confirmed is entirely `build_job_result`'s job
        # now, so behavior is unchanged for callers with no confirmer
        # configured - see `build_job_result`'s "no confirmer" fallback.
        best = candidates[0]
        return PendingLLMConfirmation(
            dp=dp,
            candidates_json=candidates_json,
            query_text=dp.product_name or dp.normalized_name or "",
            candidate_text=best.master_product.product_name or "",
            top_score=top_score,
            best_master_product_id=best.master_product.id,
        )

    return "no_match", {
        **_destination_json(dp),
        "candidates": candidates_json,
        "score": top_score,
    }


def _build_auto_match_confirmer() -> LLMAutoMatchConfirmer | None:
    """Built once per run (not per destination row - constructing an LLM
    client per row would be wasteful and, for a real API client, slow).
    Returns None whenever `enable_llm_auto_match` is off or the selected
    provider has no API key configured, so a misconfigured or disabled
    setup silently falls back to ordinary manual review - never a crash,
    same rule as `build_reranker`'s own LLM wiring.

    Logs which of those three outcomes happened - this used to be
    genuinely silent, which made "why didn't this get auto-matched" hard
    to answer from `docker compose logs backend` alone. First real user
    report of this exact confusion: an obviously-correct match with a
    typo stayed in manual review with no way to tell whether the LLM
    declined it or was never actually configured/called at all.
    """
    if not settings.enable_llm_auto_match:
        print("[llm_auto_match] disabled (ENABLE_LLM_AUTO_MATCH is not true) - skipping for this run")
        return None
    llm_client = _build_llm_client(
        settings.llm_reranker_provider,
        settings.anthropic_api_key,
        settings.llm_reranker_model,
        settings.gemini_api_key,
        settings.gemini_reranker_model,
    )
    if llm_client is None:
        print(
            f"[llm_auto_match] enabled but no usable client for provider "
            f"{settings.llm_reranker_provider!r} (missing API key?) - skipping for this run"
        )
        return None
    print(
        f"[llm_auto_match] active: provider={settings.llm_reranker_provider!r} "
        f"min_score={settings.medium_confidence_threshold}"
    )
    return LLMAutoMatchConfirmer(llm_client, min_score=settings.medium_confidence_threshold)


def build_job_result(
    db: Session, index: CatalogSearchIndex, destination_upload_id: str, job_id: str,
    master_upload_id: str | None = None,
) -> dict:
    destination_products = (
        db.query(DestinationProduct)
        .filter(DestinationProduct.upload_id == destination_upload_id)
        .order_by(DestinationProduct.source_row)
        .all()
    )
    total = len(destination_products)
    buckets: dict[str, list[dict]] = {"auto_matched": [], "needs_review": [], "no_match": []}
    confirmer = _build_auto_match_confirmer()
    pending: list[PendingLLMConfirmation] = []

    # Phase 1: classify every row. Exact/hybrid-threshold matches and clear
    # no-matches are decided immediately; anything that might need an LLM
    # confirmation is collected into `pending` instead of calling the LLM
    # per row (HANDOFF.md section 10.2/10.4). Reserve the last 10% of
    # progress for phase 2 below when there's a confirmer to run - a run
    # with no confirmer configured needs it since phase 2 is a no-op then.
    #
    # A row whose status is no longer "pending" already has a real decision
    # behind it - either Part A's per-click /decisions save earlier in THIS
    # run, or a genuinely earlier session being resumed (Part B, NEXT_STEPS
    # item 6). Report that decision as-is (_resume_record_for_decided)
    # instead of reclassifying - see that function's own docstring for why
    # this must never silently re-decide something a human already
    # confirmed.
    classify_weight = 0.9 if confirmer is not None else 1.0
    for i, dp in enumerate(destination_products):
        if dp.status != "pending":
            bucket, record = _resume_record_for_decided(db, dp)
            buckets[bucket].append(record)
        else:
            result = _classify_destination_product(db, index, dp, master_upload_id=master_upload_id)
            if isinstance(result, PendingLLMConfirmation):
                pending.append(result)
            else:
                bucket, record = result
                buckets[bucket].append(record)
        percent = round(((i + 1) / total) * 100 * classify_weight) if total else 100
        _update_job(job_id, current=i + 1, percent=percent)

    # Phase 2: one batched LLM pass over everything phase 1 deferred,
    # instead of one API call per row - the actual fix for a real
    # 20-request/day free-tier quota that used to be exhausted in under a
    # minute (config.py's llm_batch_size). If no confirmer is configured,
    # every pending row falls back to the same score-based split
    # `_classify_destination_product` always used, unchanged from before
    # this refactor.
    if confirmer is not None and pending:
        print(
            f"[llm_auto_match] confirming {len(pending)} candidate(s) via LLM, "
            f"batched {max(1, settings.llm_batch_size)} per call"
        )
        decisions = confirmer.confirm_batch(
            [(p.query_text, p.candidate_text, p.top_score) for p in pending]
        )
    else:
        decisions = [False] * len(pending)

    for p, confirmed in zip(pending, decisions):
        if confirmed:
            buckets["auto_matched"].append(
                {
                    **_destination_json(p.dp),
                    "candidates": p.candidates_json,
                    "selected_catalog_id": p.best_master_product_id,
                    "score": p.top_score,
                    "auto_match_source": "llm_confirmed",
                }
            )
        elif p.top_score >= settings.low_confidence_threshold:
            buckets["needs_review"].append(
                {
                    **_destination_json(p.dp),
                    "candidates": p.candidates_json,
                    "score": p.top_score,
                }
            )
        else:
            buckets["no_match"].append(
                {
                    **_destination_json(p.dp),
                    "candidates": p.candidates_json,
                    "score": p.top_score,
                }
            )

    _update_job(job_id, current=total, percent=100)

    return {
        "auto_matched": buckets["auto_matched"],
        "needs_review": buckets["needs_review"],
        "no_match": buckets["no_match"],
        "stats": {
            "auto_matched": len(buckets["auto_matched"]),
            "needs_review": len(buckets["needs_review"]),
            "no_match": len(buckets["no_match"]),
        },
    }


def list_catalog_versions(db: Session) -> list[CatalogVersion]:
    """For the wizard's "use existing catalog" picker (HANDOFF.md section 4).
    Newest first, since that is almost always the one to default to.
    """
    return db.query(CatalogVersion).order_by(CatalogVersion.created_at.desc()).all()


def _create_catalog_version(db: Session, master_upload: Upload, product_count: int) -> CatalogVersion:
    """Wraps a freshly-ingested master Upload as a reusable CatalogVersion,
    so the next run can say "use this one" instead of re-ingesting and
    re-indexing the same Excel file again.

    `product_count` is the INDEXED count (group headers excluded - see
    IndexStats.indexed_records), not the raw row count, so the number shown
    in the picker matches what will actually be searched.
    """
    version = CatalogVersion(
        name=master_upload.filename,
        source_upload_id=master_upload.id,
        product_count=product_count,
        is_active=True,
    )
    db.add(version)
    db.flush()
    return version


def list_resumable_destinations(db: Session, limit: int = 20) -> list[dict[str, Any]]:
    """Feeds the wizard's "continue a previous run" picker (Part B of the
    "no resume" fix, NEXT_STEPS.md item 6). Newest first, same ordering
    convention as `list_catalog_versions`.

    Deliberately includes destination uploads with zero pending rows too
    (not just in-progress ones) - a reviewer might want to reopen a
    finished run to re-export it, or double-check something, and filtering
    those out here would be a surprising, undocumented restriction with no
    real benefit (the wizard's own "0 remaining" count already makes a
    finished upload obvious at a glance).
    """
    uploads = (
        db.query(Upload)
        .filter(Upload.upload_type == "destination")
        .order_by(Upload.created_at.desc())
        .limit(limit)
        .all()
    )
    result = []
    for upload in uploads:
        total = (
            db.query(DestinationProduct)
            .filter(DestinationProduct.upload_id == upload.id)
            .count()
        )
        if total == 0:
            continue  # nothing to resume - an empty/failed ingest, not a real run
        pending = (
            db.query(DestinationProduct)
            .filter(DestinationProduct.upload_id == upload.id, DestinationProduct.status == "pending")
            .count()
        )
        result.append(
            {
                "upload_id": upload.id,
                "filename": upload.filename,
                "created_at": upload.created_at.isoformat() if upload.created_at else None,
                "total": total,
                "pending": pending,
                "decided": total - pending,
            }
        )
    return result


def run_matching_job(
    job_id: str,
    master_path: str | None,
    master_filename: str | None,
    destination_path: str | None,
    destination_filename: str | None,
    catalog_version_id: str | None = None,
    destination_upload_id: str | None = None,
) -> None:
    """Runs in a background thread (see the API layer). Classifies every
    destination product against a master catalog, in one of two modes:

    - `catalog_version_id` given: reuse an already-ingested catalog. This is
      the fix for HANDOFF.md section 4's "re-ingests 5,163 rows on every
      run" problem - no ingest_master() call at all, and the search index
      is pulled from the per-version cache (index_manager.py) if a previous
      run already built it, rather than rebuilt from scratch (~11s over the
      real catalog). On a cache miss (e.g. after a server restart, since the
      cache is in-memory only) it rebuilds once from that version's
      source_upload_id and re-populates the cache, so the cost is paid at
      most once per process lifetime per version, not once per run.
    - `master_path`/`master_filename` given instead: the original one-shot
      behavior - ingest the uploaded Excel fresh, build the index, and
      (new) wrap the result in a CatalogVersion so it becomes reusable via
      the first path on the *next* run.

    Similarly, exactly one of two things provides the DESTINATION side:

    - `destination_upload_id` given (Part B, "continue a previous run" -
      NEXT_STEPS.md item 6): reuse an already-ingested destination upload
      instead of ingesting a fresh file. `build_job_result` already treats
      any DestinationProduct whose `status != "pending"` as a real,
      already-made decision rather than something to reclassify (see
      `_resume_record_for_decided`) - that is what actually makes this a
      resume rather than a second, parallel re-review of the same rows.
    - `destination_path`/`destination_filename` given instead: the original
      one-shot behavior - ingest the uploaded Excel fresh.

    Exactly one of the two modes must be usable; neither ingests the
    destination file more than once, and both write the resulting
    catalog_version_id onto the job result so save() can stamp it on every
    Match created from this run (see save_results). Never lets an
    exception escape unhandled - the job is marked "error" instead, so a
    bad file or an unknown catalog_version_id can't hang the polling
    frontend forever.
    """
    db = SessionLocal()
    try:
        if destination_upload_id is not None:
            destination_upload = db.get(Upload, destination_upload_id)
            if destination_upload is None or destination_upload.upload_type != "destination":
                raise ValueError(f"Destination upload {destination_upload_id} not found")
        else:
            if not destination_path:
                raise ValueError("Provide either destination_upload_id (resume) or a destination file to ingest")
            destination_upload = ingest_destination(db, destination_path, destination_filename or "destination.xlsx", IngestionOptions())
            db.commit()

        if catalog_version_id is not None:
            catalog_version = db.get(CatalogVersion, catalog_version_id)
            if catalog_version is None:
                raise ValueError(f"Catalog version {catalog_version_id} not found")
            master_upload_id = catalog_version.source_upload_id

            index = get_cached_index_for_version(catalog_version_id)
            if index is None:
                # Cache miss - normal right after a restart, since the
                # cache is process-memory only. Scoped to just this
                # version's source upload, same reasoning as the fresh-file
                # branch below: unscoped load_master_records() would mix in
                # every other catalog ever ingested.
                records = load_master_records(db, upload_id=master_upload_id)
                index = CatalogSearchIndex()
                index.build(records, collection_name=_qdrant_collection_name(master_upload_id))
                cache_index_for_version(catalog_version_id, index)
        else:
            if not master_path:
                raise ValueError("Provide either catalog_version_id or a catalog file to ingest")

            master_upload = ingest_master(db, master_path, master_filename or "catalog.xlsx", IngestionOptions())
            db.commit()

            # Scoped to just this run's freshly-ingested master upload - not
            # all MasterProduct rows ever ingested. Real bug found in
            # production: an earlier standalone-wizard run (or a mistaken
            # file drop) left rows in master_product from a *different*
            # file; since rows are never deleted, an unscoped
            # load_master_records() here would silently mix that old
            # catalog's data into this run's index and produce bogus "exact
            # matches" against a file that was never uploaded this time.
            records = load_master_records(db, upload_id=master_upload.id)
            index = CatalogSearchIndex()
            stats = index.build(records, collection_name=_qdrant_collection_name(master_upload.id))

            catalog_version = _create_catalog_version(db, master_upload, stats.indexed_records)
            db.commit()
            cache_index_for_version(catalog_version.id, index)
            catalog_version_id = catalog_version.id
            master_upload_id = master_upload.id

        result = build_job_result(db, index, destination_upload.id, job_id, master_upload_id=master_upload_id)
        result["catalog_version_id"] = catalog_version_id
        _update_job(job_id, status="done", percent=100, result=result)
    except Exception as exc:  # noqa: BLE001 - must never crash the background thread silently
        _update_job(job_id, status="error", error=str(exc))
    finally:
        db.close()


# --- Save: persist decisions + generate the Excel export (spec section 28) -


_DECISION_TO_TYPE = {
    "авто": "auto_accepted",
    "вручную": "user_selected",
    "отклонено": "no_match",
    "без совпадения": "no_match",
}

_DECISION_TO_MATCH_TYPE = {
    "авто": "auto_accepted",
    "вручную": "user_selected",
}

# How an "авто" decision was actually reached, keyed by the
# `auto_match_source` field _classify_destination_product puts on every
# auto_matched record (see its docstring). Distinct from `match_type` above
# (which spec section 28's export cares about) - this is specifically for
# `Match.method`'s audit trail, so "an LLM confirmed this despite a
# depressed score" is never indistinguishable from "this cleared the
# hybrid threshold on its own" or "this was a verbatim exact match".
#
# Falls back to "exact_match" for a missing/unrecognized source - the
# original, pre-LLM-auto-match behavior for every "авто" row, preserved as
# the default so nothing that worked before this feature existed changes.
_AUTO_MATCH_SOURCE_TO_METHOD = {
    "exact_match": "exact_match",
    "hybrid_threshold": "auto_accepted_hybrid",
    "llm_confirmed": "llm_auto_matched",
}


def save_results(db: Session, rows: list[dict], catalog_version_id: str | None = None) -> None:
    """Persists every row's decision as real Match/Feedback rows and
    updates destination_products.status - the same tables every other tab
    reads and writes, so this tool's decisions show up in feedback-stats
    and count toward Phase 7's training-data threshold too.

    `catalog_version_id` is the one attached to the run's job result (see
    run_matching_job) and is stamped on every Match created here - this is
    the piece HANDOFF.md section 4 calls "hard to reverse": once a catalog
    is replaced and its MasterProduct rows overwritten in place, a Match
    with no catalog_version_id has no way to reconstruct what the reviewer
    actually approved. Left None if the caller doesn't have one (e.g. an
    older frontend build, or a direct API call that predates this field) -
    an honest NULL rather than a guessed value.

    `row["already_persisted"] = True` skips a row entirely (see the "no
    resume" fix, NEXT_STEPS.md item 6). Two sources produce this flag:
    a decision the wizard already sent to `/decisions` the instant it was
    made this session, and a decision resumed from a genuinely earlier
    session (`_resume_record_for_decided`, tagged `already_decided` and
    carried through to `already_persisted` by the frontend's
    `buildSaveRows`). Either way, the write already happened; doing it
    again here would insert a second, redundant Match/Feedback row every
    single time the final "Save & Export" is clicked - this is what
    actually makes calling `/decisions` per-click and `/save` at the end
    safe to do together, rather than double-writing everything.
    """
    for row in rows:
        if row.get("already_persisted"):
            continue

        destination_id = row.get("destination_id")
        dp = db.get(DestinationProduct, destination_id) if destination_id else None
        if dp is None:
            continue  # destination product no longer exists; skip rather than crash the whole save

        decision = row.get("decision", "")
        if decision == "пропущено":
            # Skipped during manual review: an explicit "decide later" that is
            # deliberately NOT persisted as a Match/Feedback and does not touch
            # the destination product's status - it stays whatever it was
            # (typically pending) so the item can be picked up again on a later
            # run rather than being silently recorded as a no-match.
            continue
        catalog_id = row.get("catalog_id")
        score = row.get("score")
        decision_type = _DECISION_TO_TYPE.get(decision, "no_match")

        if catalog_id and decision in _DECISION_TO_MATCH_TYPE:
            master_product = db.get(MasterProduct, catalog_id)
            if decision == "авто":
                method = _AUTO_MATCH_SOURCE_TO_METHOD.get(row.get("auto_match_source"), "exact_match")
            else:
                method = "standalone_wizard"
            db.add(
                Match(
                    destination_product_id=dp.id,
                    master_product_id=catalog_id,
                    confidence=score,
                    match_type=decision_type,
                    method=method,
                    is_confirmed=True,
                    catalog_version_id=catalog_version_id,
                )
            )
            db.add(
                Feedback(
                    destination_product_id=dp.id,
                    selected_master_product_id=catalog_id,
                    decision_type=decision_type,
                    candidate_data={
                        "destination_product": dp.product_name,
                        "selected_master_product": master_product.product_name if master_product else None,
                        "candidates": [{"id": catalog_id, "rank": 1, "score": score}],
                        "decision": decision_type,
                    },
                )
            )
            dp.status = "matched"
        else:
            db.add(
                Feedback(
                    destination_product_id=dp.id,
                    selected_master_product_id=None,
                    decision_type="no_match",
                    candidate_data={
                        "destination_product": dp.product_name,
                        "selected_master_product": None,
                        "candidates": [],
                        "decision": "no_match",
                    },
                )
            )
            dp.status = "no_match"

    db.flush()


# Column order follows how the result is actually read: what was
# requested, then what it was matched to, then how confident that was.
#
# Descriptions and prices are included because they are what makes the
# export verifiable after the fact - a name and a percentage alone cannot
# be checked against the catalog without reopening it. The internal UUID
# is kept, but last, since it is only needed for tracing back into the
# database and is noise to a human reading the sheet.
_EXPORT_HEADERS = [
    "Наименование (заявка)",
    "Описание (заявка)",
    "Цена (заявка)",
    "Код (каталог)",
    "Наименование (каталог)",
    "Описание (каталог)",
    "Цена (каталог)",
    "Совпадение %",
    "Тип",
    "Проверено",
    "ID строки",
]


def build_export_workbook(rows: list[dict]) -> bytes:
    """Spec section 28 (Excel Export): a new workbook with one row per
    destination product, preserving the original destination data
    identifiers and adding match columns - never overwriting/destroying
    the source structure, per spec's explicit instruction.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matching Result"
    ws.append(_EXPORT_HEADERS)

    for row in rows:
        decision = row.get("decision", "")
        if decision == "авто" and row.get("auto_match_source") == "llm_confirmed":
            # Distinct from a plain AUTO_MATCH so anyone auditing the
            # export later can immediately see which rows an LLM, not a
            # score threshold or an exact string match, put through - the
            # same traceability principle as Match.method's
            # "llm_auto_matched" (see save_results).
            match_type = "AUTO_MATCH_AI"
        else:
            match_type = {
                "авто": "AUTO_MATCH",
                "вручную": "MANUAL_MATCH",
                "отклонено": "NO_MATCH",
                "без совпадения": "NO_MATCH",
                "пропущено": "SKIPPED",
            }.get(decision, "NO_MATCH")
        # "авто" is machine-decided (not human-reviewed); "пропущено" was
        # deliberately deferred, so it counts as not-yet-reviewed too.
        reviewed = "No" if decision in ("авто", "пропущено") else "Yes"
        score = row.get("score")
        confidence_pct = f"{round(score * 100)}%" if isinstance(score, (int, float)) else ""

        ws.append([
            row.get("destination_name"),
            row.get("destination_description") or "",
            row.get("destination_price") if row.get("destination_price") is not None else "",
            row.get("catalog_code") or row.get("catalog_id") or "",
            row.get("catalog_name") or "",
            row.get("catalog_description") or "",
            row.get("catalog_price") if row.get("catalog_price") is not None else "",
            confidence_pct,
            match_type,
            reviewed,
            row.get("destination_id"),
        ])

    for i, header in enumerate(_EXPORT_HEADERS, start=1):
        column = ws.column_dimensions[get_column_letter(i)]
        # Description columns hold multi-sentence spec text; the rest are
        # short. One uniform width would either truncate the descriptions
        # or leave the code columns absurdly wide.
        if "Описание" in header:
            column.width = 60
        elif header == "ID строки":
            column.width = 38
        else:
            column.width = max(len(header) + 2, 18)

    # Freeze the header row so it stays visible while scrolling a
    # thousand-row result.
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
