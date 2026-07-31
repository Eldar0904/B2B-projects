"""Excel reading (spec section 5).

Streams rows via openpyxl in read-only mode so large master catalogs
(potentially hundreds of thousands of rows) are not loaded into memory as
a single pandas DataFrame. Also detects sheets and the header row, since
real-world files don't always have headers on row 1 and may have leading
title rows.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterator

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetSummary:
    name: str
    row_count: int
    col_count: int
    header_row_index: int  # 0-based index into the sheet's rows
    headers: list[str]


def list_sheets(path: str, max_header_scan_rows: int = 15) -> list[SheetSummary]:
    """List all sheets in a workbook with detected headers."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        summaries = []
        for name in wb.sheetnames:
            ws = wb[name]
            header_idx, headers = _detect_header_row(ws, max_header_scan_rows)
            summaries.append(
                SheetSummary(
                    name=name,
                    row_count=ws.max_row or 0,
                    col_count=ws.max_column or 0,
                    header_row_index=header_idx,
                    headers=headers,
                )
            )
        return summaries
    finally:
        wb.close()


def _detect_header_row(ws: Worksheet, max_scan: int) -> tuple[int, list[str]]:
    """Heuristic: the header row is the row (within the first `max_scan`
    rows) with the highest count of non-empty, unique, string-typed cells.
    Falls back to row 0 if nothing scores above zero.
    """
    best_idx = 0
    best_score = -1
    best_headers: list[str] = []

    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        string_cells = [c for c in row if isinstance(c, str) and c.strip()]
        unique_ratio = len(set(string_cells)) / len(string_cells) if string_cells else 0
        score = len(string_cells) * unique_ratio
        if score > best_score:
            best_score = score
            best_idx = idx
            best_headers = [str(c).strip() if c is not None else "" for c in row]

    return best_idx, best_headers


def iter_rows(path: str, sheet_name: str, header_row_index: int) -> Iterator[tuple[int, dict]]:
    """Yield (source_row_number, row_dict) for every data row after the header.

    `source_row_number` is the 1-based Excel row number, preserved for
    traceability (spec section 38: every match must be traceable back to
    the original row).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        headers: list[str] = []
        for excel_row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if excel_row_num - 1 < header_row_index:
                continue
            if excel_row_num - 1 == header_row_index:
                headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(row)]
                continue
            if all(c is None for c in row):
                continue
            row_dict = {headers[i] if i < len(headers) else f"col_{i}": v for i, v in enumerate(row)}
            yield excel_row_num, row_dict
    finally:
        wb.close()
