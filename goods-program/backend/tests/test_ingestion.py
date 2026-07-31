import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import DestinationProduct, MasterProduct, Upload
from app.services.ingestion import IngestionOptions, ingest_destination, ingest_master


@pytest.fixture
def db_session():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture
def master_xlsx(tmp_path):
    path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalog"
    ws.append(["Код", "Наименование", "Единица измерения", "Сметная цена, тенге"])
    ws.append(["521-101-0100", "Оборудования игровое", None, None])  # group header row
    ws.append(["521-101-0131-0001", "Манеж детский размерами 830х680 мм", "шт.", 8482])
    ws.append(["521-101-0131-0002", "Манеж детский размерами 840х840х680 мм", "шт.", "not a number"])
    wb.save(path)
    return str(path)


@pytest.fixture
def destination_xlsx(tmp_path):
    path = tmp_path / "destination.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "детсад"
    ws.append(["Код", "Наименование товара", "Цена с НДС, в тенге", "SUM из Кол-во"])
    ws.append(["D001", "Стол для детей регулируемый", 45000, 10])
    ws.append(["D002", "Стульчик пластиковый", 12000, 20])
    wb.save(path)
    return str(path)


def test_ingest_master_marks_group_headers_and_keeps_bad_price_row(db_session, master_xlsx):
    upload = ingest_master(db_session, master_xlsx, "master.xlsx")
    db_session.commit()

    assert upload.status == "done"
    assert upload.processed_rows == 3
    assert upload.skipped_rows == 0

    products = db_session.query(MasterProduct).order_by(MasterProduct.source_row).all()
    assert products[0].is_group_header is True
    assert products[0].price is None
    assert products[1].is_group_header is False
    assert products[1].price == 8482.0
    # unparseable price should not crash ingestion; price just ends up None
    assert products[2].price is None
    assert products[2].product_name == "Манеж детский размерами 840х840х680 мм"


def test_ingest_destination_maps_and_normalizes(db_session, destination_xlsx):
    upload = ingest_destination(db_session, destination_xlsx, "destination.xlsx")
    db_session.commit()

    assert upload.status == "done"
    assert upload.processed_rows == 2

    products = db_session.query(DestinationProduct).order_by(DestinationProduct.source_row).all()
    assert products[0].external_id == "D001"
    assert products[0].quantity == 10.0
    assert products[0].price == 45000.0
    assert products[0].normalized_name == "стол для детей регулируемый"
    # raw_data must preserve the original row for traceability
    assert products[0].raw_data["Наименование товара"] == "Стол для детей регулируемый"


def test_ingestion_never_crashes_on_completely_empty_sheet(db_session, tmp_path):
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Код", "Наименование"])
    wb.save(path)

    upload = ingest_master(db_session, str(path), "empty.xlsx")
    db_session.commit()
    assert upload.status == "done"
    assert upload.processed_rows == 0


def test_long_external_id_is_truncated_not_crashed(db_session, tmp_path):
    """Real bug found in production: this project's actual master catalog
    has section-header rows where the entire multi-line category title
    (300+ characters) sits in the "Код" column instead of a real product
    code. SQLite never enforces VARCHAR(255), so this was invisible until
    running against real PostgreSQL, which correctly rejects an oversized
    value - see ARCHITECTURE.md / ingestion.py for the fix (truncate to
    fit, keep the full original text in raw_data).
    """
    long_code = "Оборудование, мебель и инвентарь для объектов образования " * 6  # > 255 chars
    assert len(long_code) > 255

    path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Код", "Наименование", "Единица измерения", "Сметная цена, тенге"])
    ws.append([long_code, None, None, None])
    ws.append(["521-101-0131-0001", "Манеж детский", "шт.", 8482])
    wb.save(path)

    upload = ingest_master(db_session, str(path), "master.xlsx")
    db_session.commit()

    assert upload.status == "done"
    assert upload.processed_rows == 2
    assert upload.skipped_rows == 0

    products = db_session.query(MasterProduct).order_by(MasterProduct.source_row).all()
    assert len(products[0].external_id) == 255  # truncated to fit VARCHAR(255)
    assert products[0].external_id == long_code[:255]
    # the full, untruncated original value is still preserved for traceability
    assert products[0].raw_data["Код"] == long_code


def test_flush_batch_with_row_fallback_salvages_good_rows_after_bad_one():
    """Unit test for the batch-level resilience: if flushing a whole
    batch fails at the database level, the fallback commits row-by-row so
    one bad row can't take an entire batch of otherwise-good rows down
    with it.
    """
    from unittest.mock import MagicMock

    from sqlalchemy.exc import DataError

    from app.services.ingestion import _flush_batch_with_row_fallback

    db = MagicMock()
    # First flush() call (the whole-batch attempt) fails; every per-row
    # flush() after that succeeds.
    db.flush.side_effect = [DataError("stmt", {}, Exception("value too long")), None, None, None]

    batch = ["record1", "record2", "record3"]
    source_rows = [10, 11, 12]
    errors = []

    succeeded, failed = _flush_batch_with_row_fallback(db, batch, errors, source_rows)

    assert succeeded == 3
    assert failed == 0
    assert errors == []
    assert db.rollback.call_count == 1  # only the initial whole-batch failure rolled back


def test_flush_batch_with_row_fallback_isolates_the_actual_bad_row():
    from unittest.mock import MagicMock

    from sqlalchemy.exc import DataError

    from app.services.ingestion import _flush_batch_with_row_fallback

    db = MagicMock()
    bad_row_error = DataError("stmt", {}, Exception("value too long for type character varying(255)"))
    # Whole-batch flush fails, then per-row: row1 ok, row2 fails, row3 ok.
    db.flush.side_effect = [bad_row_error, None, bad_row_error, None]

    batch = ["record1", "record2", "record3"]
    source_rows = [10, 11, 12]
    errors = []

    succeeded, failed = _flush_batch_with_row_fallback(db, batch, errors, source_rows)

    assert succeeded == 2
    assert failed == 1
    assert len(errors) == 1
    assert errors[0]["row"] == 11
    assert "too long" in errors[0]["reason"]
