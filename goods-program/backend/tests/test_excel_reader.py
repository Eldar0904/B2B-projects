import openpyxl
import pytest

from app.services import excel_reader


@pytest.fixture
def sample_xlsx(tmp_path):
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Код", "Наименование", "Цена"])
    ws.append(["001", "Стол детский", 45000])
    ws.append(["002", "Стул детский", 12000])
    ws.append([None, None, None])  # blank row should be skipped
    ws.append(["003", "Шкаф", 30000])
    wb.save(path)
    return str(path)


def test_list_sheets_detects_header_row(sample_xlsx):
    sheets = excel_reader.list_sheets(sample_xlsx)
    assert len(sheets) == 1
    assert sheets[0].name == "Products"
    assert sheets[0].header_row_index == 0
    assert sheets[0].headers == ["Код", "Наименование", "Цена"]


def test_iter_rows_skips_blank_rows_and_preserves_row_numbers(sample_xlsx):
    rows = list(excel_reader.iter_rows(sample_xlsx, "Products", header_row_index=0))
    row_numbers = [r[0] for r in rows]
    assert row_numbers == [2, 3, 5]  # Excel row 4 (blank) is skipped
    assert rows[0][1]["Наименование"] == "Стол детский"
    assert rows[2][1]["Код"] == "003"


def test_header_detection_skips_leading_title_row(tmp_path):
    path = tmp_path / "with_title.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Мастер каталог продуктов 2026"])  # title row, single cell
    ws.append(["Код", "Наименование", "Цена"])
    ws.append(["001", "Стол", 1000])
    wb.save(path)

    sheets = excel_reader.list_sheets(str(path))
    assert sheets[0].header_row_index == 1
    assert sheets[0].headers == ["Код", "Наименование", "Цена"]
